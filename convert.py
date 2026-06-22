#!/usr/bin/env python3
"""Convert mixed-format data files into JSONL for ingest.py.

One row/line -> one JSON record. Handles CSV/TSV, Excel (.xlsx), delimited
text / combolists (e.g. email:password), and JSON/JSONL. Point it at a file or
a whole directory (walked recursively) and it streams JSONL to stdout, ready to
pipe into ingest.py.

    # whole folder -> hashed + loaded in one pipe
    python convert.py "C:/leaks" | python ingest.py - --email-field email --username-field username --phone-field phone

    # one Excel file to a JSONL file
    python convert.py leak.xlsx --out leak.jsonl

    # headerless combolist with a custom layout
    python convert.py combo.txt --delimiter : --columns email,password,ip

Identifier column names are canonicalized by default (mail/e-mail/login -> email,
tel/mobile -> phone, etc.) so the same ingest flags work on every file. Each
record also gets a `source_file` field for provenance (disable with --no-source-file).

Excel needs openpyxl on the host:  pip install openpyxl
(everything else is stdlib). .xls (old binary) is not supported - re-save as
.xlsx or CSV.
"""

import argparse
import csv
import json
import os
import sys

CSV_EXT = {".csv"}
TSV_EXT = {".tsv", ".tab"}
XLSX_EXT = {".xlsx"}
JSON_EXT = {".json"}
JSONL_EXT = {".jsonl", ".ndjson"}
TEXT_EXT = {".txt", ".text", ".list", ".combo", ".dat", ""}
SUPPORTED = CSV_EXT | TSV_EXT | XLSX_EXT | JSON_EXT | JSONL_EXT | TEXT_EXT

# identifier header aliases -> canonical field name
ALIASES = {
    "email": "email", "e-mail": "email", "mail": "email", "emailaddress": "email",
    "email_address": "email", "login_email": "email", "user_email": "email",
    "username": "username", "user": "username", "user_name": "username",
    "login": "username", "handle": "username", "screenname": "username",
    "screen_name": "username", "nick": "username", "nickname": "username",
    "phone": "phone", "phone_number": "phone", "phonenumber": "phone",
    "mobile": "phone", "tel": "phone", "telephone": "phone", "msisdn": "phone",
    "password": "password", "pass": "password", "passwd": "password", "pwd": "password",
}

COMBO_DELIMS = [":", "|", ";", "\t", ","]


def warn(msg):
    print(f"convert: {msg}", file=sys.stderr)


def canon_key(key, enabled):
    k = str(key).strip().lower().replace(" ", "_")
    if enabled and k in ALIASES:
        return ALIASES[k]
    return k or "field"


def clean_record(raw, canonical, source_file, add_source):
    """Drop empty values, canonicalize keys, add provenance."""
    rec = {}
    for k, v in raw.items():
        if v is None:
            continue
        s = v if isinstance(v, (int, float, bool)) else str(v).strip()
        if s == "":
            continue
        rec[canon_key(k, canonical)] = s
    if not rec:
        return None
    if add_source and "source_file" not in rec:
        rec["source_file"] = source_file
    return rec


def rows_from_csv(path, delimiter, columns, has_header_opt):
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        if delimiter is None:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except csv.Error:
                delimiter = ","
        if has_header_opt is None:
            try:
                has_header = csv.Sniffer().has_header(sample)
            except csv.Error:
                has_header = True
        else:
            has_header = has_header_opt
        reader = csv.reader(fh, delimiter=delimiter)
        header = None
        for i, row in enumerate(reader):
            if not any(c.strip() for c in row):
                continue
            if header is None:
                if columns:
                    header = columns
                    if has_header and not has_header_opt:
                        pass  # ambiguous; trust explicit columns, treat row as data
                elif has_header:
                    header = [c.strip() or f"col{j+1}" for j, c in enumerate(row)]
                    continue
                else:
                    header = [f"col{j+1}" for j in range(len(row))]
            yield dict(zip(header, row))


def rows_from_xlsx(path, sheet):
    try:
        from openpyxl import load_workbook
    except ImportError:
        warn(f"skipping {path}: Excel support needs openpyxl  (pip install openpyxl), "
             f"or re-save the file as CSV")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = [wb[sheet]] if sheet else wb.worksheets
    for ws in sheets:
        it = ws.iter_rows(values_only=True)
        header = None
        for row in it:
            if row is None or all(c is None for c in row):
                continue
            if header is None:
                header = [str(c).strip() if c is not None else f"col{j+1}"
                          for j, c in enumerate(row)]
                continue
            yield {header[j] if j < len(header) else f"col{j+1}": row[j]
                   for j in range(len(row))}


def rows_from_text(path, delimiter, columns):
    with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
        first = ""
        for line in fh:
            if line.strip():
                first = line.strip()
                break
        if not first:
            return
        delim = delimiter
        if delim is None:
            delim = next((d for d in COMBO_DELIMS if d in first), None)
        cols = columns
        if delim is None:
            cols = cols or ["line"]
        elif not cols:
            n = len(first.split(delim))
            if n == 2:
                cols = ["email", "password"]
            else:
                warn(f"{path}: {n} fields per line on '{delim}' - pass --columns "
                     f"to name them (defaulting to col1..col{n})")
                cols = [f"col{i+1}" for i in range(n)]
        fh.seek(0)
        for line in fh:
            line = line.rstrip("\n").rstrip("\r")
            if not line.strip():
                continue
            if delim is None:
                yield {cols[0]: line}
            else:
                parts = line.split(delim, len(cols) - 1)
                yield {cols[i]: parts[i] for i in range(min(len(cols), len(parts)))}


def rows_from_json(path):
    with open(path, "r", encoding="utf-8-sig") as fh:
        try:
            data = json.load(fh)
        except json.JSONDecodeError as e:
            warn(f"skipping {path}: invalid JSON ({e})")
            return
    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                yield item
    elif isinstance(data, dict):
        yield data


def rows_from_jsonl(path):
    with open(path, "r", encoding="utf-8-sig") as fh:
        for n, line in enumerate(fh, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                warn(f"{path} line {n}: malformed JSON; skipped ({e})")
                continue
            if isinstance(obj, dict):
                yield obj


def rows_for_file(path, args):
    ext = os.path.splitext(path)[1].lower()
    if ext in XLSX_EXT:
        return rows_from_xlsx(path, args.sheet)
    if ext in JSON_EXT:
        return rows_from_json(path)
    if ext in JSONL_EXT:
        return rows_from_jsonl(path)
    if ext in CSV_EXT:
        return rows_from_csv(path, args.delimiter, args.columns, args.has_header)
    if ext in TSV_EXT:
        return rows_from_csv(path, args.delimiter or "\t", args.columns, args.has_header)
    return rows_from_text(path, args.delimiter, args.columns)


def iter_files(paths):
    for p in paths:
        if os.path.isdir(p):
            for root, _, files in os.walk(p):
                for name in sorted(files):
                    fp = os.path.join(root, name)
                    if os.path.splitext(name)[1].lower() in SUPPORTED:
                        yield fp
        elif os.path.isfile(p):
            yield p
        else:
            warn(f"not found: {p}")


def main():
    ap = argparse.ArgumentParser(description="Convert mixed-format files to JSONL.")
    ap.add_argument("paths", nargs="+", help="File(s) or directory(ies) to convert")
    ap.add_argument("--out", default=None, help="Write JSONL here (default: stdout)")
    ap.add_argument("--delimiter", default=None,
                    help="Field delimiter for CSV/text (default: auto-detect)")
    ap.add_argument("--columns", default=None,
                    help="Comma-separated column names for headerless CSV/text")
    ap.add_argument("--has-header", dest="has_header", action="store_true", default=None,
                    help="Force: first CSV row IS a header")
    ap.add_argument("--no-header", dest="has_header", action="store_false",
                    help="Force: CSV/text has NO header row")
    ap.add_argument("--sheet", default=None, help="Excel: only this sheet (default: all)")
    ap.add_argument("--no-canonical", dest="canonical", action="store_false", default=True,
                    help="Do not rename identifier columns to email/username/phone")
    ap.add_argument("--no-source-file", dest="source_file", action="store_false", default=True,
                    help="Do not add a source_file provenance field")
    args = ap.parse_args()
    if args.columns:
        args.columns = [c.strip() for c in args.columns.split(",")]

    out = open(args.out, "w", encoding="utf-8") if args.out else sys.stdout
    files = total = written = 0
    try:
        for path in iter_files(args.paths):
            files += 1
            src = os.path.splitext(os.path.basename(path))[0]
            for raw in rows_for_file(path, args):
                total += 1
                rec = clean_record(raw, args.canonical, src, args.source_file)
                if rec is None:
                    continue
                out.write(json.dumps(rec, ensure_ascii=False) + "\n")
                written += 1
    finally:
        if args.out:
            out.close()
    warn(f"done: {files} file(s), {total} rows read, {written} records written")


if __name__ == "__main__":
    main()
